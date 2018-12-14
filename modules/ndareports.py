#!/usr/bin/env python

# ---AUTHOR---
# Name: Matt Cross
# Email: routeallthings@gmail.com
#
# ndareports.py

# Import Native
import os
import re
from decimal import *

# Import NDA
from downloadfile import *

try:
	from openpyxl import load_workbook
	from openpyxl import workbook
	from openpyxl import Workbook
	from openpyxl.styles import Font, NamedStyle
except ImportError:
	openpyxlinstallstatus = raw_input ('openpyxl module is missing, would you like to automatically install? (Y/N): ')
	if 'y' in openpyxlinstallstatus.lower():
		os.system('python -m pip install openpyxl')
		from openpyxl import load_workbook
		from openpyxl import workbook
		from openpyxl import Workbook
		from openpyxl.styles import Font, NamedStyle
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of openpyxl. Please install manually and retry'
		print 'Exiting in 5 seconds'
		time.sleep(5)
		sys.exit()		
# PyPiWin32
try:
	import win32com.client
except ImportError:
	win32cominstallstatus = raw_input ('PyPiWin32 module is missing, would you like to automatically install? (Y/N): ')
	if 'y' in win32cominstallstatus.lower():
		os.system('python -m pip install pypiwin32')
		os.system('python -m pip install pywin32')
		print 'You need to restart the script after installing win32com'
		print 'Exiting in 5 seconds'
		time.sleep(5)
		sys.exit()
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of PyPiWin32. Please install manually and retry'
		print 'Exiting in 5 seconds'
		time.sleep(5)
		sys.exit()		
#
try:
	import fileinput
except ImportError:
	requestsinstallstatus = fullpath = raw_input ('FileInput module is missing, would you like to automatically install? (Y/N): ')
	if 'Y' in requestsinstallstatus or 'y' in requestsinstallstatus or 'yes' in requestsinstallstatus or 'Yes' in requestsinstallstatus or 'YES' in requestsinstallstatus:
		os.system('python -m pip install FileInput')
		import FileInput
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of FileInput. Please install manually and retry'
		sys.exit()

# Get root path and add macdb library and template library
import inspect, os.path, sys
filename = inspect.getframeinfo(inspect.currentframe()).filename
reprootpath = os.path.dirname(os.path.abspath(filename))
# MAC OUI Library
reprpath,lastfolder = os.path.split(reprootpath)
lastfolder = 'macdb'
macdbpath = os.path.join(reprpath,lastfolder)
# Template Library
reprpath,lastfolder = os.path.split(reprootpath)
lastfolder = 'templates'
templatepath = os.path.join(reprpath,lastfolder)		


# Report Style
HeaderFont = Font(bold=True)
HeaderFont.size = 12
HeaderStyle = NamedStyle(name='BoldHeader')
HeaderStyle.font = HeaderFont
	
### Full Inventory ###
def fullinventoryreport(fullinventorylist,exportlocation):
	# Start Report
	wb = Workbook()
	dest_filename = 'Full-Inventory.xlsx'
	dest_path = exportlocation + '\\' + dest_filename
	ws1 = wb.active
	ws1.title = "Device Inventory"
	ws1.append(['Hostname','Product ID','Serial Number','Stack Number','Manufacture Date','Version','Description'])
	# Continue on with work
	startrow = 2
	for row in fullinventorylist:
		if row.get('Description').lower().endswith('chassis') == True or 'k9' in row.get('Description').lower():
			# Attempt to find the age of the device
			try:
				age_base = 1996
				age_year = int(row.get('Serial Number')[3:5])
				age_week = (row.get('Serial Number')[5:7])
				age_year_manufactured = age_base + age_year
				age_manufactured = datetime.datetime.strptime(str(age_year_manufactured) + '-W' + age_week.encode('utf-8') + '-0', '%Y-W%W-%w')
				age_manufactured = '{:%B %d, %Y}'.format(age_manufactured)
			except:
				age_manufactured = ''
			# Add to workbook
			ws1['A' + str(startrow)] = row.get('Hostname')
			ws1['B' + str(startrow)] = row.get('Product ID')
			ws1['C' + str(startrow)] = row.get('Serial Number')
			ws1['D' + str(startrow)] = row.get('Stack Number')
			ws1['E' + str(startrow)] = age_manufactured
			ws1['F' + str(startrow)] = row.get('Version')
			ws1['G' + str(startrow)] = row.get('Description')
			startrow = startrow + 1
	# Start Secondary Tab
	ws2 = wb.create_sheet(title="Module Inventory")
	# Continue on with work
	ws2.append(['Hostname','Product ID','Serial Number','Description'])
	startrow = 2
	for row in fullinventorylist:
		ws2['A' + str(startrow)] = row.get('Hostname')
		ws2['B' + str(startrow)] = row.get('Product ID')
		ws2['C' + str(startrow)] = row.get('Serial Number')
		ws2['D' + str(startrow)] = row.get('Description')
		startrow = startrow + 1
	wb.add_named_style(HeaderStyle)
	# Set styles on header row
	for cell in ws1["1:1"]:
		cell.style = 'BoldHeader'
	for cell in ws2["1:1"]:
		cell.style = 'BoldHeader'
	# Set Column Width
	for col in ws1.columns:
		max_length = 0
		column = col[0].column # Get the column name
		for cell in col:
			try: # Necessary to avoid error on empty cells
				if len(str(cell.value)) > max_length:
					max_length = len(cell.value)
			except:
				pass
		adjusted_width = (max_length + 2) * 1.2
		ws1.column_dimensions[column].width = adjusted_width
	for col in ws2.columns:
		max_length = 0
		column = col[0].column # Get the column name
		for cell in col:
			try: # Necessary to avoid error on empty cells
				if len(str(cell.value)) > max_length:
					max_length = len(cell.value)
			except:
				pass
		adjusted_width = (max_length + 2) * 1.2
		ws2.column_dimensions[column].width = adjusted_width
	# Save File
	try:
		wb.save(filename = dest_path)
	except:
		print 'Error creating the report: ' + dest_path + '. File might be currently in use.'
		return
	# Sorting
	try:
		# Tab 1
		excel = win32com.client.Dispatch("Excel.Application")
		wb = excel.Workbooks.Open(dest_path)
		ws = wb.Worksheets('Device Inventory')
		ws.Range('A2:G50000').Sort(Key1=ws.Range('A1'), Order1=1, Orientation=1)
		ws.Range('A1:G1').AutoFilter(1)
		wb.Save()
		# Tab 2
		wb = excel.Workbooks.Open(dest_path)
		ws = wb.Worksheets('Module Inventory')
		ws.Range('A2:D50000').Sort(Key1=ws.Range('A1'), Order1=1, Orientation=1)
		ws.Range('A1:D1').AutoFilter(1)
		wb.Save()
		excel.Application.Quit()
	except:
		# Throw no error
		pass
	print 'Successfully created Full Inventory Report'
	
def arpmacreport(iparptablelist,ipmactablelist,mactablelist,exportlocation):
	tempfilelist = []
	maclookupdburl = "http://standards-oui.ieee.org/oui.txt"
	wb = Workbook()
	dest_filename = 'ARP-MAC-Report.xlsx'
	dest_path = exportlocation + '\\' + dest_filename
	ws1 = wb.active
	# Continue on with work
	ws1.title = "ARP Report"
	ws1.append(['IP Address','MAC','Manufacturer','Source Device','Inteface','MAC Count on Interface'])
	# Create ARP report by looking for closest hop interface
	skiparpreport = 0
	# Preload MAC DB
	try:
		maclookupdb = []
		maclookupfilename = 'oui.txt'
		maclookupdbo = os.path.join(macdbpath,maclookupfilename)
		with open(maclookupdbo, 'r') as maclookupdbop:
			maclookupdb_unfiltered = maclookupdbop.readlines()
		for line in maclookupdb_unfiltered:
			if 'hex' in line:
				maclookupdb_dict = {}
				maclookupreg = re.search('(\S+)\s+\S+\s+(.*)',line)
				macwithouthiphens = maclookupreg.group(1).replace('-','')
				maclookupdb_dict['mac'] = macwithouthiphens
				maclookupdb_dict['company'] = maclookupreg.group(2)
				maclookupdb.append(maclookupdb_dict)
		skipmac = 0
	except Exception as e:
		skipmac = 1
		print 'Could not load MAC database. Error is ' + str(e)
	# Start processing data
	for row in ipmactablelist:
		tempdict = {}
		if row.get('IP Address') == 'Incomplete':
			continue
		tempdict['IP Address'] = row.get('IP Address')
		tempdict['MAC'] = row.get('MAC')
		maccompany = ''
		mac_company_mac = str(row.get('MAC')[0:7].replace('.','')).upper()
		# Get a vendor mac address and add to the table
		try:
			macsearch = filter(lambda x: x['mac'] == mac_company_mac, maclookupdb)
			if macsearch == []:
				maccompany = 'Unknown'
			else:
				for dict in macsearch:
					maccompany = dict['company']
		except:
			maccompany = 'Unknown'
		tempdict['MAC Manufacturer'] = maccompany
		if '-' in row.get('Age'):
			tempdict['Source Device'] = row.get('Hostname')
			tempdict['Interface'] = row.get('Interface')
			tempdict['MAC Count on Interface'] = 1
		else:
			macint = ''
			machost = ''
			# Find the lowest count interface in the list that matches the mac address
			maccount = 100000
			for temprow in mactablelist:
				if temprow.get('Count') <= maccount and temprow.get('MAC') == row.get('MAC'):
					maccount = temprow.get('Count')
					macint = temprow.get('Interface')
					machost = temprow.get('Hostname')
			# Bug Fix - If it could not calculate, give the number a non-integer unknown
			if maccount == 100000:
				maccount = 'Unknown'
			tempdict['Source Device'] = machost
			tempdict['Interface'] = macint
			tempdict['MAC Count on Interface'] = maccount
		iparptablelist.append(tempdict)
	# Create the actual ARP report
	startrow = 2
	for row in iparptablelist:
		ws1['A' + str(startrow)] = row.get('IP Address')
		ws1['B' + str(startrow)] = row.get('MAC')
		ws1['C' + str(startrow)] = row.get('MAC Manufacturer')
		ws1['D' + str(startrow)] = row.get('Source Device')
		ws1['E' + str(startrow)] = row.get('Interface')
		ws1['F' + str(startrow)] = row.get('MAC Count on Interface')
		startrow = startrow + 1
	# Change worksheet
	ws2 = wb.create_sheet(title="MAC Report")
	# Continue on with work
	ws2.append(['Hostname','MAC','Manufacturer','VLAN','Interface'])
	startrow = 2
	for row in mactablelist:
		# Get Manufacturer of MAC
		maccompany = ''
		mac_company_mac = str(row.get('MAC')[0:7].replace('.','')).upper()
		# Get a vendor mac address and add to the table
		try:
			macsearch = filter(lambda x: x['mac'] == mac_company_mac, maclookupdb)
			if macsearch == []:
				maccompany = 'Unknown'
			else:
				for dict in macsearch:
					maccompany = dict['company']
		except:
			maccompany = 'Unknown'
		# Append to Workbook
		ws2['A' + str(startrow)] = row.get('Hostname')
		ws2['B' + str(startrow)] = row.get('MAC')
		ws2['C' + str(startrow)] = maccompany
		ws2['D' + str(startrow)] = row.get('VLAN')
		ws2['E' + str(startrow)] = row.get('Interface')
		startrow = startrow + 1
	wb.add_named_style(HeaderStyle)
	# Set styles on header row
	for cell in ws1["1:1"]:
		cell.style = 'BoldHeader'
	for cell in ws2["1:1"]:
		cell.style = 'BoldHeader'
	# Set Column Width
	for col in ws1.columns:
		max_length = 0
		column = col[0].column # Get the column name
		for cell in col:
			try: # Necessary to avoid error on empty cells
				if len(str(cell.value)) > max_length:
					max_length = len(cell.value)
			except:
				pass
		adjusted_width = (max_length + 2) * 1.2
		ws1.column_dimensions[column].width = adjusted_width
	for col in ws2.columns:
		max_length = 0
		column = col[0].column # Get the column name
		for cell in col:
			try: # Necessary to avoid error on empty cells
				if len(str(cell.value)) > max_length:
					max_length = len(cell.value)
			except:
				pass
		adjusted_width = (max_length + 2) * 1.2
		ws2.column_dimensions[column].width = adjusted_width
	# Save workbook
	try:
		wb.save(filename = dest_path)
	except:
		print 'Error creating the report: ' + dest_path + '. File might be currently in use.'
		return
	# Sorting
	try:
		# Tab 1
		excel = win32com.client.Dispatch("Excel.Application")
		wb = excel.Workbooks.Open(dest_path)
		ws = wb.Worksheets('ARP Report')
		ws.Range('A2:F50000').Sort(Key1=ws.Range('A1'), Order1=1, Orientation=1)
		ws.Range('A1:F1').AutoFilter(1)
		wb.Save()
		# Tab 2
		wb = excel.Workbooks.Open(dest_path)
		ws = wb.Worksheets('MAC Report')
		ws.Range('A2:E50000').Sort(Key1=ws.Range('A1'), Order1=1, Orientation=1)
		ws.Range('A1:E1').AutoFilter(1)
		wb.Save()
		excel.Application.Quit()
	except:
		# Throw no error
		pass
	print 'Successfully created ARP/MAC Report'
	# Cleanup
	for file in tempfilelist:
		try:
			os.remove(file)
		except:
			pass
	
	
def interfacereport (l2interfacelist,l3interfacelist,poeinterfacelist,exportlocation):
	wb = Workbook()
	dest_filename = 'Interface-Report.xlsx'
	dest_path = exportlocation + '\\' + dest_filename
	ws1 = wb.active
	# Continue on with work
	ws1.title = "Interface Overview"
	ws1.append(['Hostname','100Mb','1Gb','10gb','40gb','100gb','POE'])
	startrow = 2
	# Populate Device Names
	l2devicenames = []
	for row in l2interfacelist:
		dupdetect = 0	
		for device in l2devicenames:
			if row.get('Hostname') == device:
				dupdetect = 1
		if dupdetect == 0:
			l2devicenames.append(row.get('Hostname'))
	# Count Interfaces
	for row in l2devicenames:
		int_hostname = row
		faint = 0
		geint = 0
		tengeint = 0
		fortygeint = 0
		hundredgeint = 0
		poeint = 0
		for subrow in l2interfacelist:
			if '100BaseTX' in subrow.get('Type') and int_hostname == subrow.get('Hostname'):
				faint = faint + 1
			if '1000BaseTX' in subrow.get('Type') and int_hostname == subrow.get('Hostname'):
				geint = geint + 1
			if '10000BaseTX' in subrow.get('Type') and int_hostname == subrow.get('Hostname'):
				tengeint = tengeint + 1
			if '10/40GB' in subrow.get('Type') and int_hostname == subrow.get('Hostname'):
				fortygeint = fortygeint + 1
			if '10/25/50/100' in subrow.get('Type') and int_hostname == subrow.get('Hostname'):
				hundredgeint = hundredgeint + 1
			if '10G' in subrow.get('Type') and int_hostname == subrow.get('Hostname'):
				tengeint = tengeint + 1
			if '10/40/100' in subrow.get('Type') and int_hostname == subrow.get('Hostname'):
				hundredgeint = hundredgeint + 1
			for subrow1 in poeinterfacelist:
				if subrow1.get('Hostname') == int_hostname and subrow.get('Hostname') == int_hostname:
					if subrow.get('Interface') == subrow1.get('Interface') and Decimal(subrow1.get('Power Usage')) > 0:
						poeint = poeint + 1
		# Get Hostname and populate
		ws1['A' + str(startrow)] = int_hostname
		ws1['B' + str(startrow)] = faint
		ws1['C' + str(startrow)] = geint
		ws1['D' + str(startrow)] = tengeint
		ws1['E' + str(startrow)] = fortygeint
		ws1['F' + str(startrow)] = hundredgeint
		ws1['G' + str(startrow)] = poeint
		startrow = startrow + 1
	try:
		ws2 = wb.create_sheet(title="L2 Interfaces")
		# Continue on with work
		ws2.append(['Hostname','Interface','Type','Status','Speed','Duplex','VLAN','POE'])
		startrow = 2
		for row in l2interfacelist:
			# Get POE interfaces and combine
			for subrow in poeinterfacelist:
				if row.get('Interface') == subrow.get('Interface') and Decimal(subrow.get('Power Usage')) > 0 and row.get('Hostname') == subrow.get('Hostname'):
					poeint = 'Yes'
				else:
					poeint = 'No'
			ws2['A' + str(startrow)] = row.get('Hostname')
			ws2['B' + str(startrow)] = row.get('Interface')
			ws2['C' + str(startrow)] = row.get('Type')
			ws2['D' + str(startrow)] = row.get('Status')
			ws2['E' + str(startrow)] = row.get('Speed')
			ws2['F' + str(startrow)] = row.get('Duplex')
			ws2['G' + str(startrow)] = row.get('VLAN').encode('utf-8')
			ws2['H' + str(startrow)] = poeint
			startrow = startrow + 1
	except Exception as e:
		print 'Error creating L2 Interface Report. Error is ' + str(e)	
	try:
		ws3 = wb.create_sheet(title="L3 Interfaces")
		# Continue on with work
		ws3.append(['Hostname','Interface','IP Address'])
		startrow = 2
		for row in l3interfacelist:
			if not 'unassigned' in row.get('IP Address').lower():
				ws3['A' + str(startrow)] = row.get('Hostname')
				ws3['B' + str(startrow)] = row.get('Interface')
				ws3['C' + str(startrow)] = row.get('IP Address')
				startrow = startrow + 1
	except Exception as e:
		print 'Error creating L3 Interface Report. Error is ' + str(e)	
	wb.add_named_style(HeaderStyle)
	# Set styles on header row
	for cell in ws1["1:1"]:
		cell.style = 'BoldHeader'
	for cell in ws2["1:1"]:
		cell.style = 'BoldHeader'
	for cell in ws3["1:1"]:
		cell.style = 'BoldHeader'
	# Set Column Width
	for col in ws1.columns:
		max_length = 0
		column = col[0].column # Get the column name
		for cell in col:
			try: # Necessary to avoid error on empty cells
				if len(str(cell.value)) > max_length:
					max_length = len(cell.value)
			except:
				pass
		adjusted_width = (max_length + 2) * 1.2
		ws1.column_dimensions[column].width = adjusted_width
	for col in ws2.columns:
		max_length = 0
		column = col[0].column # Get the column name
		for cell in col:
			try: # Necessary to avoid error on empty cells
				if len(str(cell.value)) > max_length:
					max_length = len(cell.value)
			except:
				pass
		adjusted_width = (max_length + 2) * 1.2
		ws2.column_dimensions[column].width = adjusted_width
	for col in ws3.columns:
		max_length = 0
		column = col[0].column # Get the column name
		for cell in col:
			try: # Necessary to avoid error on empty cells
				if len(str(cell.value)) > max_length:
					max_length = len(cell.value)
			except:
				pass
		adjusted_width = (max_length + 2) * 1.2
		ws3.column_dimensions[column].width = adjusted_width
	# Save File
	try:
		wb.save(filename = dest_path)
	except:
		print 'Error creating the report: ' + dest_path + '. File might be currently in use.'
		return
	# Sorting
	try:
		# Tab 1
		excel = win32com.client.Dispatch("Excel.Application")
		wb = excel.Workbooks.Open(dest_path)
		ws = wb.Worksheets('Interface Overview')
		ws.Range('A2:G50000').Sort(Key1=ws.Range('A1'), Order1=1, Orientation=1)
		ws.Range('A1:G1').AutoFilter(1)
		wb.Save()
		# Tab 2
		wb = excel.Workbooks.Open(dest_path)
		ws = wb.Worksheets('L2 Interfaces')
		ws.Range('A2:H50000').Sort(Key1=ws.Range('A1'), Order1=1, Orientation=1)
		ws.Range('A1:H1').AutoFilter(1)
		wb.Save()
		# Tab 3
		wb = excel.Workbooks.Open(dest_path)
		ws = wb.Worksheets('L3 Interfaces')
		ws.Range('A2:C50000').Sort(Key1=ws.Range('A1'), Order1=1, Orientation=1)
		ws.Range('A1:C1').AutoFilter(1)
		wb.Save()
		excel.Application.Quit()
	except:
		# Throw no error
		pass
	print 'Successfully created Interface Report'

def poereport(poeinterfacelist,exportlocation):
	wb = Workbook()
	dest_filename = 'POE-Report.xlsx'
	dest_path = exportlocation + '\\' + dest_filename
	ws1 = wb.active
	# Continue on with work
	ws1.title = "POE Interfaces"
	ws1.append(['Hostname','Interface','Admin Status','Operational Status','Power Usage','Device Name','Device Class','Max POE Capability'])
	startrow = 2
	for row in poeinterfacelist:
		if not row.get('Power Usage') == '0.0':
			# Add to workbook
			ws1['A' + str(startrow)] = row.get('Hostname')
			ws1['B' + str(startrow)] = row.get('Interface')
			ws1['C' + str(startrow)] = row.get('Admin Status')
			ws1['D' + str(startrow)] = row.get('Up/Down')
			ws1['E' + str(startrow)] = row.get('Power Usage')
			ws1['F' + str(startrow)] = row.get('Device Name')
			ws1['G' + str(startrow)] = row.get('Device Class')
			ws1['H' + str(startrow)] = row.get('Max POE Capability')
			startrow = startrow + 1
	wb.add_named_style(HeaderStyle)
	# Set styles on header row
	for cell in ws1["1:1"]:
		cell.style = 'BoldHeader'
	# Set Column Width
	for col in ws1.columns:
		max_length = 0
		column = col[0].column # Get the column name
		for cell in col:
			try: # Necessary to avoid error on empty cells
				if len(str(cell.value)) > max_length:
					max_length = len(cell.value)
			except:
				pass
		adjusted_width = (max_length + 2) * 1.2
		ws1.column_dimensions[column].width = adjusted_width
	# Save File
	try:
		wb.save(filename = dest_path)
	except:
		print 'Error creating the report: ' + dest_path + '. File might be currently in use.'
		return
	# Sorting
	try:
		excel = win32com.client.Dispatch("Excel.Application")
		wb = excel.Workbooks.Open(dest_path)
		ws = wb.Worksheets('POE Interfaces')
		ws.Range('A2:H50000').Sort(Key1=ws.Range('A1'), Order1=1, Orientation=1)
		ws.Range('A1:H1').AutoFilter(1)
		wb.Save()
		excel.Application.Quit()
	except:
		# Throw no error
		pass
	print 'Successfully created POE Report'
	
def healthcheckreport(healthchecklist,exportlocation):
	wb = Workbook()
	dest_filename = 'Health-Check-Report.xlsx'
	dest_path = exportlocation + '\\' + dest_filename
	ws1 = wb.active
	# Continue on with work
	ws1.title = "Health Check"
	ws1.append(['Hostname','Error','Description'])
	startrow = 2
	for row in healthchecklist:
		ws1['A' + str(startrow)] = row.get('Hostname')
		ws1['B' + str(startrow)] = row.get('Error')
		ws1['C' + str(startrow)] = row.get('Description')
		startrow = startrow + 1	
	wb.add_named_style(HeaderStyle)
	# Set styles on header row
	for cell in ws1["1:1"]:
		cell.style = 'BoldHeader'
	# Set Column Width
	for col in ws1.columns:
		max_length = 0
		column = col[0].column # Get the column name
		for cell in col:
			try: # Necessary to avoid error on empty cells
				if len(str(cell.value)) > max_length:
					max_length = len(cell.value)
			except:
				pass
		adjusted_width = (max_length + 2) * 1.2
		ws1.column_dimensions[column].width = adjusted_width
	# Save File
	try:
		wb.save(filename = dest_path)
	except:
		print 'Error creating the report: ' + dest_path + '. File might be currently in use.'
		return
	# Sorting
	try:
		excel = win32com.client.Dispatch("Excel.Application")
		wb = excel.Workbooks.Open(dest_path)
		ws = wb.Worksheets('Health Check')
		ws.Range('A2:C50000').Sort(Key1=ws.Range('A1'), Order1=1, Orientation=1)
		ws.Range('A1:C1').AutoFilter(1)
		wb.Save()
		excel.Application.Quit()
	except:
		# Throw no error
		pass
	print 'Successfully created Health Check Report'