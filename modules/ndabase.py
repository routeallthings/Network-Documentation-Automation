#!/usr/bin/env python

# ---AUTHOR---
# Name: Matt Cross
# Email: routeallthings@gmail.com
#
# ndabase.py

##### Native Imports #####
import re
import getpass
import os
import unicodedata
import csv
import threading
import time
import sys
import json
import logging
import datetime
import socket
import struct
import json

# Paramiko Fixes

paramiko_logger = logging.getLogger('paramiko.transport')
if not paramiko_logger.handlers:
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(
        logging.Formatter('%(asctime)s | %(levelname)-8s| PARAMIKO: '
                      '%(lineno)03d@%(module)-10s| %(message)s')
    )
paramiko_logger.addHandler(console_handler)

##### Non-Native Imports #####

# Used for pulling information for MAC table
try:
	import requests
except ImportError:
	requestsinstallstatus = raw_input ('request module is missing, would you like to automatically install? (Y/N): ')
	if 'y' in requestsinstallstatus.lower():
		os.system('python -m pip install requests')
		import requests
	else:
		print "You selected an option other than yes. Please be aware that this script requires the use of requests. Please install manually and retry"
		print 'Exiting in 5 seconds'
		time.sleep(5)
		sys.exit()

# Used because its the best way to connect into the switches to pull out information (besides SNMP)
try:
	import netmiko
	from netmiko import ConnectHandler
	from paramiko.ssh_exception import SSHException 
	from netmiko.ssh_exception import NetMikoTimeoutException
except ImportError:
	netmikoinstallstatus = fullpath = raw_input ('Netmiko module is missing, would you like to automatically install? (Y/N): ')
	if 'y' in netmikoinstallstatus.lower():
		os.system('python -m pip install netmiko')
		os.system('python -m pip install utils')
		import netmiko
		from netmiko import ConnectHandler
		from paramiko.ssh_exception import SSHException 
		from netmiko.ssh_exception import NetMikoTimeoutException
	else:
		print "You selected an option other than yes. Please be aware that this script requires the use of netmiko. Please install manually and retry"
		print 'Exiting in 5 seconds'
		time.sleep(5)
		sys.exit()

# Used to parse through SSH output to get relevant data into tables for comparison and output
try:
	import textfsm
except ImportError:
	textfsminstallstatus = raw_input ('textfsm module is missing, would you like to automatically install? (Y/N): ')
	if 'y' in textfsminstallstatus.lower():
		os.system('python -m pip install textfsm')
		import textfsm
	else:
		print "You selected an option other than yes. Please be aware that this script requires the use of textfsm. Please install manually and retry"
		print 'Exiting in 5 seconds'
		time.sleep(5)
		sys.exit()

# XLSX import
try:
	from openpyxl import load_workbook
	from openpyxl import workbook
	from openpyxl import Workbook
	from openpyxl.styles import Font, NamedStyle
except ImportError:
	openpyxlinstallstatus = raw_input ('openpyxl module is missing, would you like to automatically install? (Y/N): ')
	if 'y' in openpyxlinstallstatus.lower():
		os.system('python -m pip install openpyxl')
		from openpyxl import load_workbook
		from openpyxl import workbook
		from openpyxl import Workbook
		from openpyxl.styles import Font, NamedStyle
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of openpyxl. Please install manually and retry'
		print 'Exiting in 5 seconds'
		time.sleep(5)
		sys.exit()

# XLSX import
try:
	import fileinput
except ImportError:
	fileinputnstallstatus = raw_input ('FileInput module is missing, would you like to automatically install? (Y/N): ')
	if 'y' in fileinputnstallstatus.lower():
		os.system('python -m pip install FileInput')
		import FileInput
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of FileInput. Please install manually and retry'
		print 'Exiting in 5 seconds'
		time.sleep(5)
		sys.exit()

# Darth-Veitcher Module https://github.com/darth-veitcher/xlhelper
# XLSX Import

from pprint import pprint
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from collections import OrderedDict
try:
	import xlhelper
except ImportError:
	xlhelperinstallstatus = raw_input ('xlhelper module is missing, would you like to automatically install? (Y/N): ')
	if 'y' in xlhelperinstallstatus.lower():
		os.system('python -m pip install git+https://github.com/routeallthings/xlhelper.git')
		import xlhelper
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of xlhelper. Please install manually and retry'
		print 'Exiting in 5 seconds'
		time.sleep(5)
		sys.exit()		

# PyPiWin32
try:
	import win32com.client
except ImportError:
	win32cominstallstatus = raw_input ('PyPiWin32 module is missing, would you like to automatically install? (Y/N): ')
	if 'y' in win32cominstallstatus.lower():
		os.system('python -m pip install pypiwin32')
		os.system('python -m pip install pywin32')
		print 'You need to restart the script after installing win32com'
		print 'Exiting in 5 seconds'
		time.sleep(5)
		sys.exit()
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of PyPiWin32. Please install manually and retry'
		print 'Exiting in 5 seconds'
		time.sleep(5)
		sys.exit()	
		
# MNETSUITE (https://github.com/MJL85/mnet/tree/master/mnetsuite)---- THIS HAS BEEN MODIFIED IN ROUTEALLTHINGS FORK
# CDP/LLDP Discovery and Mapping method

try:
	import mnetsuite_routeallthings
except ImportError:
	mnetinstallstatus = raw_input ('mnetsuite module (routeallthings variant) is missing, would you like to automatically install? (Y/N): ')
	if 'y' in mnetinstallstatus.lower():
		os.system('python -m pip install git+git://github.com/routeallthings/mnet_routeallthings.git')
		import mnetsuite_routeallthings
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of mnetsuite. Please install manually and retry'
		print 'Exiting in 5 seconds'
		time.sleep(5)
		sys.exit()